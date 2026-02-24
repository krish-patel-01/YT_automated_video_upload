"""Metadata handling for video uploads via Excel queue file."""

import logging
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import xlwings as xw

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Excel column definitions
# ---------------------------------------------------------------------------
COL_FILENAME      = "video_filename"   # A  – video file name (e.g. clip.mp4)
COL_TITLE         = "title"            # B  – video title
COL_DESCRIPTION   = "description"      # C  – video description
COL_GEN_TAGS      = "generate_tags"    # D  – "yes" / "no"
COL_GEN_TAGS_OUT  = "generated_tags"   # E  – written back by the system
COL_UPLOAD        = "upload"           # F  – "yes" / "no"  (triggers upload)
COL_STATUS        = "status"           # G  – Uploaded / Failed / (blank)

EXPECTED_COLUMNS = [
    COL_FILENAME,
    COL_TITLE,
    COL_DESCRIPTION,
    COL_GEN_TAGS,
    COL_GEN_TAGS_OUT,
    COL_UPLOAD,
    COL_STATUS,
]

# Column widths (characters)
_COL_WIDTHS = {
    "A": 40,   # video_filename
    "B": 60,   # title
    "C": 80,   # description
    "D": 15,   # generate_tags
    "E": 70,   # generated_tags
    "F": 12,   # upload
    "G": 15,   # status
}


class VideoMetadata:
    """Container for video metadata."""

    def __init__(
        self,
        title: str,
        description: str = "",
        tags: list[str] = None,
        category_id: str = "25",
        privacy_status: str = "public",
        made_for_kids: bool = False,
    ):
        self.title = title
        self.description = description
        self.tags = tags or []
        self.category_id = category_id
        self.privacy_status = privacy_status
        self.made_for_kids = made_for_kids


class MetadataHandler:
    """Load video metadata from an Excel queue file."""

    def __init__(
        self,
        excel_file: str,
        default_description: str,
        default_tags: list[str],
        default_category_id: str = "25",
        default_privacy_status: str = "public",
        default_made_for_kids: bool = False,
    ):
        """Initialize metadata handler.

        Args:
            excel_file: Path to the Excel upload queue file
            default_description: Default description for news videos
            default_tags: Default tags for news videos
            default_category_id: YouTube category ID (25 = News & Politics)
            default_privacy_status: Default privacy status
            default_made_for_kids: Whether videos are made for kids
        """
        self.excel_file = excel_file
        self.default_description = default_description
        self.default_tags = default_tags
        self.default_category_id = default_category_id
        self.default_privacy_status = default_privacy_status
        self.default_made_for_kids = default_made_for_kids

        self._ensure_excel_file_exists()

    # ------------------------------------------------------------------
    # Excel file management
    # ------------------------------------------------------------------

    def _ensure_excel_file_exists(self):
        """Create the Excel queue file with headers if it does not exist."""
        if os.path.exists(self.excel_file):
            return

        logger.info(f"Creating Excel queue file: {self.excel_file}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Upload Queue"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(EXPECTED_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Column widths
        for letter, width in _COL_WIDTHS.items():
            ws.column_dimensions[letter].width = width

        ws.row_dimensions[1].height = 30
        wb.save(self.excel_file)
        logger.info(f"Excel queue file created: {self.excel_file}")

    def _load_workbook(self):
        """Load the Excel workbook."""
        try:
            return openpyxl.load_workbook(self.excel_file)
        except Exception as e:
            logger.error(f"Failed to load Excel file {self.excel_file}: {e}")
            return None

    def _get_column_map(self, ws) -> Dict[str, int]:
        """Return a mapping of lowercase column name → 1-based column index."""
        col_map = {}
        for cell in ws[1]:
            if cell.value:
                col_map[str(cell.value).strip().lower()] = cell.column
        return col_map
    def _write_cell_live(self, row: int, col: int, value: str):
        """Write a single cell into the Excel file via xlwings COM interface.

        Works whether the workbook is currently open in Excel or not.
        If Excel has the file open the edit appears live without any save
        dialog.  If the file is not open it is opened invisibly, written,
        saved, and closed automatically.

        Args:
            row: 1-based row index
            col: 1-based column index
            value: Value to write into the cell
        """
        abs_path = str(Path(self.excel_file).resolve())
        abs_path_lower = abs_path.lower()

        # Check whether any running Excel instance already has this file open
        opened_new_app = False
        xl_app = None
        wb = None
        try:
            for app in xw.apps:
                for book in app.books:
                    if book.fullname.lower() == abs_path_lower:
                        wb = book
                        break
                if wb:
                    break

            if wb is None:
                # File not open in Excel – launch a hidden instance
                xl_app = xw.App(visible=False, add_book=False)
                opened_new_app = True
                wb = xl_app.books.open(abs_path)

            ws = wb.sheets[0]
            ws.cells(row, col).value = value
            wb.save()
            logger.debug(f"xlwings: wrote row={row} col={col} value={value!r}")

        except Exception as e:
            logger.error(f"xlwings write failed (row={row}, col={col}): {e}")
            raise
        finally:
            if opened_new_app and xl_app is not None:
                try:
                    xl_app.quit()
                except Exception:
                    pass
    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def load_metadata(self, video_path: str) -> "VideoMetadata":
        """Load metadata for a video by looking it up in the Excel queue.

        Priority for tags: generated_tags column > default_tags.
        Falls back to default values if the video is not listed.

        Args:
            video_path: Absolute path to the video file

        Returns:
            VideoMetadata instance
        """
        filename = Path(video_path).name
        wb = self._load_workbook()

        if wb is not None:
            ws = wb.active
            col_map = self._get_column_map(ws)

            missing = [c for c in [COL_FILENAME, COL_TITLE] if c not in col_map]
            if missing:
                logger.warning(
                    f"Excel file is missing columns: {missing}. Using default metadata."
                )
            else:
                fn_col   = col_map[COL_FILENAME]
                title_col = col_map[COL_TITLE]
                desc_col  = col_map.get(COL_DESCRIPTION)
                gen_tags_col = col_map.get(COL_GEN_TAGS_OUT)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_filename = row[fn_col - 1]
                    if row_filename and str(row_filename).strip() == filename:
                        title = (
                            str(row[title_col - 1]).strip()
                            if row[title_col - 1]
                            else ""
                        )
                        description = ""
                        if desc_col and row[desc_col - 1]:
                            description = str(row[desc_col - 1]).strip()

                        if not title:
                            title = Path(video_path).stem
                        if not description:
                            description = self.default_description

                        # Use generated tags if available; otherwise defaults
                        tags = self.default_tags.copy()
                        if gen_tags_col and row[gen_tags_col - 1]:
                            raw = str(row[gen_tags_col - 1]).strip()
                            parsed = [t.strip() for t in raw.split(",") if t.strip()]
                            if parsed:
                                tags = parsed

                        logger.info(
                            f"Loaded Excel metadata for '{filename}': title='{title}'"
                        )
                        return VideoMetadata(
                            title=title,
                            description=description,
                            tags=tags,
                            category_id=self.default_category_id,
                            privacy_status=self.default_privacy_status,
                            made_for_kids=self.default_made_for_kids,
                        )

                logger.warning(
                    f"'{filename}' not found in Excel queue. Using default metadata."
                )

        return self._default_metadata(video_path)

    def mark_as_uploaded(self, video_path: str):
        """Mark the video row in Excel with status 'Uploaded'."""
        self._set_status(Path(video_path).name, "Uploaded")

    def mark_as_failed(self, video_path: str):
        """Mark the video row in Excel with status 'Failed'."""
        self._set_status(Path(video_path).name, "Failed")

    def _set_status(self, filename: str, status: str):
        """Write status into the Excel row for the given filename via xlwings.

        Compares using just the basename so the Excel cell may contain either
        a plain filename or a full path.
        """
        target_name = Path(filename).name.lower()
        wb_read = self._load_workbook()
        if wb_read is None:
            return
        ws = wb_read.active
        col_map = self._get_column_map(ws)
        fn_col     = col_map.get(COL_FILENAME)
        status_col = col_map.get(COL_STATUS)
        if fn_col is None or status_col is None:
            return

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cell_val = str(row[fn_col - 1] or "").strip()
            if Path(cell_val).name.lower() == target_name:
                try:
                    self._write_cell_live(row_idx, status_col, status)
                    logger.info(f"Set status '{status}' for '{filename}' in Excel (live).")
                except Exception as e:
                    logger.warning(f"Could not update status for '{filename}': {e}")
                break

    # ------------------------------------------------------------------
    # Tag generation polling
    # ------------------------------------------------------------------

    def process_tag_generation(self, tag_generator) -> int:
        """Scan the Excel queue and generate tags for rows that request it.

        Looks for rows where:
          - generate_tags == "yes"
          - generated_tags is blank (not yet generated)

        Calls tag_generator.generate_tags(title, description) for each,
        writes the result back to the generated_tags cell, and saves the
        workbook.

        Args:
            tag_generator: TagGenerator instance

        Returns:
            Number of rows processed
        """
        wb = self._load_workbook()
        if wb is None:
            return 0

        ws = wb.active
        col_map = self._get_column_map(ws)

        fn_col        = col_map.get(COL_FILENAME)
        title_col     = col_map.get(COL_TITLE)
        desc_col      = col_map.get(COL_DESCRIPTION)
        gen_flag_col  = col_map.get(COL_GEN_TAGS)
        gen_out_col   = col_map.get(COL_GEN_TAGS_OUT)

        if not all([fn_col, title_col, gen_flag_col, gen_out_col]):
            logger.warning("Excel is missing tag-generation columns – skipping.")
            return 0

        processed = 0
        rows_to_write: List[Tuple[int, int, str, str]] = []  # (row, col, value, filename)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            flag_val = row[gen_flag_col - 1].value
            out_val  = row[gen_out_col - 1].value

            if not flag_val or str(flag_val).strip().lower() != "yes":
                continue
            if out_val and str(out_val).strip():
                continue  # Already persisted

            filename    = str(row[fn_col - 1].value or "").strip()
            title       = str(row[title_col - 1].value or "").strip() or filename
            desc_cell   = row[desc_col - 1].value if desc_col else None
            description = str(desc_cell or "").strip()

            logger.info(f"Generating tags for '{filename}'...")
            tags = tag_generator.generate_tags(title, description)

            if tags:
                rows_to_write.append((row_idx, gen_out_col, ", ".join(tags), filename))
                logger.info(f"Tags generated for '{filename}': {tags}")
            else:
                logger.warning(f"No tags generated for '{filename}'.")

        # Write every generated result directly into the live workbook via COM
        for row_num, col_num, tag_str, filename in rows_to_write:
            try:
                self._write_cell_live(row_num, col_num, tag_str)
                processed += 1
                logger.info(f"Tags written live to Excel for '{filename}'.")
            except Exception as e:
                logger.error(f"Failed to write tags for '{filename}': {e}")

        return processed

    # ------------------------------------------------------------------
    # Upload queue polling
    # ------------------------------------------------------------------

    def get_pending_uploads(
        self, watch_directories: List[str]
    ) -> List[Tuple[str, "VideoMetadata"]]:
        """Return a list of (video_path, VideoMetadata) for rows ready to upload.

        A row is ready when:
          - upload == "yes"
          - status is blank / not "Uploaded" / not "Failed"
          - The video file actually exists in one of the watch directories

        Args:
            watch_directories: List of folder paths to search for video files

        Returns:
            List of (absolute_video_path, VideoMetadata) tuples
        """
        wb = self._load_workbook()
        if wb is None:
            return []

        ws   = wb.active
        col_map = self._get_column_map(ws)

        fn_col       = col_map.get(COL_FILENAME)
        title_col    = col_map.get(COL_TITLE)
        desc_col     = col_map.get(COL_DESCRIPTION)
        gen_out_col  = col_map.get(COL_GEN_TAGS_OUT)
        upload_col   = col_map.get(COL_UPLOAD)
        status_col   = col_map.get(COL_STATUS)

        if not all([fn_col, upload_col]):
            logger.warning("Excel is missing required columns – skipping upload scan.")
            return []

        results: List[Tuple[str, VideoMetadata]] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            upload_val = row[upload_col - 1]
            if not upload_val or str(upload_val).strip().lower() != "yes":
                continue

            # Skip already-done rows
            if status_col:
                status_val = str(row[status_col - 1] or "").strip().lower()
                if status_val in ("uploaded", "failed"):
                    continue

            raw_entry = str(row[fn_col - 1] or "").strip()
            if not raw_entry:
                continue
            # Support both plain filenames and full paths in the Excel cell
            filename_only = Path(raw_entry).name

            # Locate the actual video file using just the basename
            video_path = self._find_video(filename_only, watch_directories)
            if video_path is None:
                logger.debug(
                    f"'{filename_only}' marked for upload but not found in watch dirs yet."
                )
                continue

            # Build metadata
            title = str(row[title_col - 1] or "").strip() if title_col else ""
            if not title:
                title = Path(filename_only).stem

            description = self.default_description
            if desc_col and row[desc_col - 1]:
                description = str(row[desc_col - 1]).strip() or self.default_description

            tags = self.default_tags.copy()
            if gen_out_col and row[gen_out_col - 1]:
                raw = str(row[gen_out_col - 1]).strip()
                parsed = [t.strip() for t in raw.split(",") if t.strip()]
                if parsed:
                    tags = parsed

            metadata = VideoMetadata(
                title=title,
                description=description,
                tags=tags,
                category_id=self.default_category_id,
                privacy_status=self.default_privacy_status,
                made_for_kids=self.default_made_for_kids,
            )
            results.append((video_path, metadata))

        return results

    @staticmethod
    def _find_video(filename: str, watch_directories: List[str]) -> Optional[str]:
        """Search watch directories for a video file by name."""
        for directory in watch_directories:
            candidate = os.path.join(directory, filename)
            if os.path.isfile(candidate):
                return candidate
        return None

    def validate_metadata(self, metadata: VideoMetadata) -> tuple[bool, list[str]]:
        """Validate video metadata.

        Args:
            metadata: VideoMetadata to validate

        Returns:
            Tuple of (is_valid, list of error messages)
        """
        errors = []

        if not metadata.title or not metadata.title.strip():
            errors.append("Title is required")
        elif len(metadata.title) > 100:
            errors.append("Title exceeds 100 characters")

        if len(metadata.description) > 5000:
            errors.append("Description exceeds 5000 characters")

        if len(metadata.tags) > 500:
            errors.append("Too many tags (max 500)")

        valid_privacy_statuses = ["public", "private", "unlisted"]
        if metadata.privacy_status not in valid_privacy_statuses:
            errors.append(
                f"Invalid privacy status. Must be one of: {valid_privacy_statuses}"
            )

        return len(errors) == 0, errors

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _default_metadata(self, video_path: str) -> VideoMetadata:
        """Generate default metadata for a video."""
        return VideoMetadata(
            title=Path(video_path).stem,
            description=self.default_description,
            tags=self.default_tags.copy(),
            category_id=self.default_category_id,
            privacy_status=self.default_privacy_status,
            made_for_kids=self.default_made_for_kids,
        )
